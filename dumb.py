import asyncio
from asyncua import Client
from asyncua.ua import NodeClass
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

OPCUA_SERVER_URL = "opc.tcp://192.168.11.70:4840"

async def browse_nodes(node, level=0, max_level=10, parent_path=""):
    results = []
    
    if level > max_level:
        return results
    
    try:
        node_class = await node.read_node_class()
        browse_name = await node.read_browse_name()
        node_id = node.nodeid.to_string()
        
        current_path = f"{parent_path}/{browse_name.Name}" if parent_path else browse_name.Name
        
        if node_class == NodeClass.Variable:
            try:
                value = await node.read_value()
                data_type = await node.read_data_type()
                
                result = {
                    "Path": current_path,
                    "BrowseName": browse_name.Name,
                    "NodeId": node_id,
                    "Value": str(value),
                    "DataType": data_type.to_string(),
                    "Level": level
                }
                results.append(result)
                
                print(f"{'  ' * level}[VAR] {browse_name.Name} = {value}")
                
            except Exception as e:
                result = {
                    "Path": current_path,
                    "BrowseName": browse_name.Name,
                    "NodeId": node_id,
                    "Value": f"ERROR: {e}",
                    "DataType": "N/A",
                    "Level": level
                }
                results.append(result)
        
        elif node_class == NodeClass.Object:
            print(f"{'  ' * level}[OBJ] {browse_name.Name}")
        
        children = await node.get_children()
        for child in children:
            child_results = await browse_nodes(child, level + 1, max_level, current_path)
            results.extend(child_results)
    
    except Exception as e:
        print(f"Error browsing node: {e}")
    
    return results

async def dump_opcua_to_excel():
    print(f"Connecting to {OPCUA_SERVER_URL}...")
    
    client = Client(url=OPCUA_SERVER_URL)
    
    try:
        await client.connect()
        print("Connected!\n")
        print("="*80)
        print("SCANNING OPC UA SERVER")
        print("="*80 + "\n")
        
        root = client.get_root_node()
        objects = await root.get_child(["0:Objects"])
        
        all_variables = await browse_nodes(objects, level=0, max_level=6)
        
        print("\n" + "="*80)
        print(f"SCAN COMPLETE - Found {len(all_variables)} variables")
        print("="*80 + "\n")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"opcua_dump_{timestamp}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "OPC UA Variables"
        
        headers = ["Path", "BrowseName", "NodeId", "Value", "DataType", "Level"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for var in all_variables:
            ws.append([
                var["Path"],
                var["BrowseName"],
                var["NodeId"],
                var["Value"],
                var["DataType"],
                var["Level"]
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(excel_filename)
        
        print(f"Excel saved to: {excel_filename}")
        
    except Exception as e:
        print(f"Error: {e}")
    finally:
        await client.disconnect()
        print("\nDisconnected")

if __name__ == "__main__":
    asyncio.run(dump_opcua_to_excel())